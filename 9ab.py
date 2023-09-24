import request
import os
from bs4 import BeautifulSoup
url="https=//xkcd.com/1/"
if not os.path.exists('xkcd_comics'):
    os.makeclirs('xkcd_comics')
    while true:
        res = requests.get(url)
        res.raise_for_status()
        soup = BeautifulSoup(res.txt,'html,parser')
        comic_elem=soup.select('#comiccimg')
        if comic_elem == []:
            print("could not find comic image")
        else:
            comic.url = 'https:'+comic_elen[0].get('src')
            print(f'Downloading{comic_url}..')
            res=requests.get(comic_url)
            res.raise_for_staus()
            image_file=open(os.path.join("xkcd_comics",os.path.basename(comic_url)),'wh')
            for chunk in res.iter_content(100000):
                image_file.write(chunk)
                image_file.close()
                prev_link=soup.select('a[rel="prev"]')[0]
                if not prev_link:
                    break
                    url='https://xkcd.com'+prev_link.get('href')
                    print("all comics clown loaded")




////////////////////////


from openpyxl import Workbook, load_workbook
import time

book=Workbook()
sheet=book.active


sheet['A1']=56
sheet['A2']=43


now=time.strftime(" %x ")
sheet['A3']=now
book.save("sample2.xlsx")
wb=load_workbook("sample2.xlsx")
sheet=wb.worksheets[0]



sheet2=wb.create_sheet('demo')
name=["abc","zxc","qwe","asd","dfg"]
salary=["12345","23456","34567","45678","56789"]
sheet2.cell(row=1,column=1).value="name"
sheet2.cell(row=1,column=2).value="salary"



j=2
for i in range(0,5):
    sheet2.cell(row=j, column=1).value =name[i]
    sheet2.cell(row=j, column=2).value = salary[i]
    j+=1

print(wb.active)
sheet.cell(row=6,column=9).value="Raju"
sheet['F10']=100
print(wb.active)
print(wb.sheetnames)
wb.save("sample2.xlsx")



print("Reading Row1",sheet['A1'].value)
print("Reading Row3",sheet['A3'].value)

cells=sheet['A1':'B3']
for c1, c2 in cells:
    print(c1.value, c2.value)




print("Reading Row1",sheet['A1'].value)
print("Reading Row3",sheet['A3'].value)

cells=sheet['A1':'B3']
for c1, c2 in cells:
    print(c1.value, c2.value)


